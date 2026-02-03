
from PyQt6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QDialog, QComboBox,
    QSizePolicy, QMenu, QStackedWidget, QListWidget
)
from PyQt6 import QtGui
from PyQt6.QtCore import Qt, QTimer, QDateTime,QPointF
from backend.mcc_backend import Mcc172Backend
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from pages.AnalysisParameters import AnalysisParameter,DisplayPreferences,InputChannelsDialog
from pages.settings_manager import load_settings, save_settings
from matplotlib.backends.backend_qt5 import NavigationToolbar2QT as NavigationToolbar
import pyqtgraph as pg
from PyQt6.QtGui import QIcon
from PyQt6.QtWidgets import QApplication
from PyQt6.QtWidgets import QMessageBox
from PyQt6.QtGui import QImage
import pandas as pd
import openpyxl
from PyQt6.QtCore import QEvent
from pages.settings_server import run_flask_background
from math import sqrt
from functools import partial

from pages.tachometer import TachometerReader
from matplotlib.figure import Figure
import os, csv, time

 
import numpy as np
from PyQt6.QtCore import QTimer, QDateTime
import pyqtgraph as pg

class TouchViewBox(pg.ViewBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setMouseEnabled(x=False, y=False)  # Disable mouse panning
        self.setAcceptTouchEvents(True) #enabing the touch object by using graphicQobject class
    #pinch to zoom of two fingers logic and also scaling done in x and y axis
        self._last_dist = None
        self._last_center = None
        self._is_pinching = False
        self._touch_start_time = 0

        self.cursor_enabled = False
        self.cursors = {}
        self.enableAutoRange(enable=True)
        self.user_has_zoomed = False
        self.manual_zoom_range = None

    def touchEvent(self, ev):
        # Skip zooming if cursor is being dragged
        if self.cursor_enabled and any(getattr(cursor, 'moving', False) for cursor in self.cursors.values()):
            super().touchEvent(ev)
            return

        

        touch_points = ev.touchPoints()
        num_touches = len(touch_points)
        if len(touch_points) == 2:
            self._is_pinching = True
            tp1,tp2 = touch_points[0], touch_points[1]


            p1 = tp1.pos()
            p2 = tp2.pos()
            
            dx = p2.x() - p1.x()
            dy = p2.y() - p1.y()
            current_dist = sqrt(dx*dx + dy*dy)
            center_x = (p1.x() + p2.x()) /2
            center_y = (p1.y() + p2.y()) /2
            center_point = QPointF(center_x, center_y)

            try:
                center_view = self.mapSceneToView(center_point)
            except:
                self._last_dist = None
                return

            if self._last_dist is not None and self._last_dist > 10:
                scale_factor = current_dist /self._last_dist

                scale_factor = max(0.5, min(2.0, scale_factor))

                self.scaleBy((1.0/scale_factor, 1.0/scale_factor), center = center_view)

                self.enableAutoRange(enable=False)
                self.user_has_zoomed = True
                self.manual_zoom_range = (self.viewRange()[0][:], self.viewRange()[1][:])

            self._last_dist = current_dist
            self._last_center = center_view

            ev.accept()
            return
        elif num_touches == 1:
            tp = touch_points[0]

            if tp.state() == tp.TouchPointState.TouchPointReleased:
                self._last_dist = None
                self._last_center = None
                self._is_pinching = False

                super().touchEvent(ev)
                return

            elif tp.state() == tp.TouchPointState.TouchPointPressed:
                self._touch_start_time = time.time()
                self._is_pinching = False
                super().touchEvent(ev)
                return

            elif tp.state() == tp.TouchPointState.TouchPointMoved:
                super().touchEvent(ev)
                return

            else:
                self._last_dist = None
                self._last_center = None
                self._is_pinching = False
                super().touchEvent(ev)

    def reset_zoom(self):
        self.user_has_zoomed = False
        self.manual_zoom_range = None
        self.enableAutoRange(enable=True)

class DraggableCursor(pg.InfiniteLine):

    def __init__(self, parent_plot, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.parent_plot = parent_plot
        self._is_moving = False
        self._initial_range = None

    def isMoving(self):
        return self._is_moving

    def mouseDragEvent(self, ev):
        if ev.button() != Qt.MouseButton.LeftButton:
            return
        if ev.isStart():
            self._is_moving = True
            vb = self.parent_plot.getViewBox()
            self._initial_range = (vb.viewRange()[0][:], vb.viewRange()[1][:])
        super().mouseDragEvent(ev)

        if self._initial_range:
            vb = self.parent_plot.getViewBox()
            vb.setRange(xRange=self._initial_range[0] ,yRange = self._initial_range[1], padding = 0)

        if ev.isFinish():
            self._is_moving = False
            self._initial_range = None

class WaveformPage(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.recording = False

        self.recordings_dir =  "recordings"
        self.raw_data_dir = "raw_data"
        self.csv_file = None
        self.video_writer = None
        self.frame_count = 0
        os.makedirs(self.recordings_dir,exist_ok = True)
        self.session_folder = ""
        try:
            run_flask_background()
        except Exception as e:
            print(f"flask server failed to start:{e}")

        for _ in range(5):
            try:
                break
            except Exception:
                time.sleep(0.05)
        
        #saving the settings into the backend json
        stored =  load_settings()

        self.selected_quantity  = stored.get("selected_quantity", "Acceleration")
        self.selected_fmax_hz   = stored.get("selected_fmax_hz", 500.0)
        self.selected_fmin_hz   = stored.get("selected_fmin_hz", 1.0)
        self.top_channel        = stored.get("top_channel", 0)
        self.bottom_channel     = stored.get("bottom_channel", 1)
        self.trace_mode_default = stored.get("trace_mode_index", 0)
        self.input_channels_data = stored.get("input_channels", None)
        self.buffer_size = stored.get("buffer_size", 8192)
        saved_sample_rate = stored.get("sample_rate", None)
        if saved_sample_rate is None:

            # Calculate from saved fmax if sample_rate not in JSON
            saved_sample_rate = int(self.selected_fmax_hz * 2.56)
       

        print("selected fmax:",self.selected_fmax_hz)
        self.daq = Mcc172Backend(buffer_size = self.buffer_size,sample_rate=saved_sample_rate)
        self.daq.setup()
        self.update_interval_ms = int((self.buffer_size / self.daq.actual_rate) * 1000 * 0.85)

        print(f"Buffer size: {self.buffer_size}, Actual sample rate: {self.daq.actual_rate}, Update interval: {self.update_interval_ms} ms")
      
        self.zoom_enabled = False
        self.cursor_enabled   = False
        self.cursors          = {}
        self.delta_labels     = {}  
        self.cursor_value_labels = {}
        self.locked_ranges = {}

        self.input_channels_dialog = InputChannelsDialog()
        if self.input_channels_data:
            self.input_channels_dialog.data = self.input_channels_data
        
        self.setup_ui()

        top_sens, bot_sens = self.get_selected_channel_sensitivities()
        self.top_sens = top_sens
        self.bot_sens = bot_sens
        self.display_settings   = stored.get("display_settings", {
            "Acceleration Spectrum Type": "RMS",
            "Acceleration Engineering Unit": "g",
            "Velocity Spectrum Type": "Peak",
            "Velocity Engineering Unit": "mm/s",

            "Displacement Spectrum Type": "RMS",
            "Displacement Engineering Unit": "µm",
        })
      
        self.selected_unit = "g"  # default unit for Acceleratio
        # self.tach = TachometerReader()
        # self.tach.rpm_updated.connect(self.update_rpm_display)

        
        self.grabGesture(Qt.GestureType.PinchGesture)

        self.start_clock()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)

        

    def setup_ui(self):
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)

        # Top Navigation Bar
        nav_bar = QHBoxLayout()
        back_btn = QPushButton("⬅")
        back_btn.setStyleSheet("""
            background-color : #444;
            color: white;
            font-weight: bold;
            padding: 6px 14px;
            border-radius: 6px;
            """)
        back_btn.clicked.connect(self.go_back_to_dashboard)
        title_label = QLabel("Waveform & Spectrum")
        title_label.setStyleSheet("font-size: 1.5rem; font-weight: bold;")
        title_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.time_label = QLabel()
        self.time_label.setStyleSheet("font-weight: bold; font-size: 1rem;")
        nav_bar.addWidget(back_btn)
        nav_bar.addWidget(title_label)
        self.record_btn = QPushButton("Record")
        self.record_btn.setIcon(QIcon("assets/record_start.png"))
        self.record_btn.setStyleSheet("""
            background-color: #dc3545;
            color: white;
            font-weight: bold;
            font-size: 1rem;
            padding: 8px;
            border-radius: 8px;
        """)
        self.record_btn.setCheckable(True)
        self.record_btn.clicked.connect(self.toggle_recording)
        nav_bar.addWidget(self.record_btn)

        self.snapshot_btn = QPushButton()
        self.snapshot_btn.setIcon(QIcon("assets/image.png"))
        self.snapshot_btn.setToolTip("Take Snapshot")
        self.snapshot_btn.clicked.connect(self.take_snapshot)
        nav_bar.addWidget(self.snapshot_btn)


        nav_bar.addWidget(self.time_label)
        self.main_layout.addLayout(nav_bar)
        self.cursor_info_label = QLabel("Cursor Info: ")
        self.cursor_info_label.setStyleSheet("font-size: 1rem; font-weight: bold; color: #003366; padding: 5px;")
        self.main_layout.addWidget(self.cursor_info_label)


        
        # Views
        self.stacked_views = QStackedWidget()
        self.default_view = self.build_readings_waveform_view()
        self.dual_view = self.build_waveform_spectrum_view()
        self.dual_waveform_view = self.build_waveform_waveform_view()
        self.dual_spectrum_view = self.build_spectrum_spectrum_view()
        self.readings_view = self.build_reading_reading_view()
        self.readings_spectrum = self.build_readings_spectrum_view()

        self.stacked_views.addWidget(self.default_view)
        self.stacked_views.addWidget(self.dual_view)
        self.stacked_views.addWidget(self.dual_waveform_view)
        self.stacked_views.addWidget(self.dual_spectrum_view)
        self.stacked_views.addWidget(self.readings_view)
        self.stacked_views.addWidget(self.readings_spectrum)

        self.main_layout.addWidget(self.stacked_views)

        self.stacked_views.setCurrentIndex(self.trace_mode_default)

        # Bottom Button Bar
        bottom_buttons = QHBoxLayout()
        bottom_buttons.setSpacing(10)
        bottom_buttons.setContentsMargins(10, 10, 10, 10)

        for label in ["Traces", "Param", "Control", "Auto", "Cursor"]:
            if label == "Traces":
                self.traces_button = QPushButton(label)
                self.traces_button.setMinimumHeight(40)
                self.traces_button.setMinimumWidth(100)
                self.traces_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
                self.traces_button.setStyleSheet("""
                    background-color: #17a2b8;
                    color: white;
                    font-weight: bold;
                    font-size: 1rem;
                    padding: 8px;
                    border-radius: 8px;
                """)
                self.traces_menu = QMenu()
                self.traces_menu.setStyleSheet("""
                    QMenu {
                        font-size: 1rem;
                        padding: 10px;
                        background-color: #f0f0f0;
                    }
                    QMenu::item {
                        padding: 10px 20px;
                        margin: 5px 0;
                    }
                    QMenu::item:selected {
                        background-color: #17a2b8;
                        color: white;
                    }
                """)
                self.traces_menu.addAction("Readings + Waveform", lambda: self.switch_trace_mode(0))
                self.traces_menu.addAction("Waveform + Spectrum", lambda: self.switch_trace_mode(1))
                self.traces_menu.addAction("Waveform + Waveform", lambda: self.switch_trace_mode(2))
                self.traces_menu.addAction("Spectrum + spectrum", lambda: self.switch_trace_mode(3))
                self.traces_menu.addAction("Readings + Readings", lambda: self.switch_trace_mode(4))
                self.traces_menu.addAction("Readings + Spectrum", lambda: self.switch_trace_mode(5))
                self.traces_menu.addSeparator()
                self.traces_menu.addAction("Trace Window Settings", self.open_trace_window_settings)
                self.traces_button.setMenu(self.traces_menu)
                bottom_buttons.addWidget(self.traces_button)
            elif label == "Param":
                self.params_button = QPushButton("Param")
                self.params_button.setMinimumHeight(40)
                self.params_button.setMinimumWidth(100)
                self.params_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
                self.params_button.setStyleSheet("""
                    background-color: #17a2b8;
                    color: black;
                    font-weight: bold;
                    font-size: 1rem;
                    padding: 8px;
                    border-radius: 8px;
                """)
                self.params_Menu = QMenu(self.params_button)
                self.params_Menu.setStyleSheet("""
                    QMenu {
                        font-size: 1rem;
                        padding: 10px;
                        background-color: #f8f9fa;
                    }
                    QMenu::item {
                        padding: 10px 20px;
                        margin: 5px 0;
                    }
                    QMenu::item:selected {
                        background-color: #17a2b8;
                        color: white;
                    }
                """)
                self.params_Menu.addAction("Analysis Parameters", self.analysis_parameter)
                self.params_Menu.addAction("Input channels", lambda: self.on_param_selected("Input Channels"))
                self.params_Menu.addAction("Output channels", lambda: self.on_param_selected("Output Channels"))
                self.params_Menu.addAction("Tachometer", lambda: self.on_param_selected("Tachometer"))
                self.params_Menu.addAction("Display Preferences", lambda: self.on_param_selected("Display Preferences"))
                self.params_Menu.addAction("view Live Signals", lambda: self.on_param_selected("view Live Signals"))
                self.params_Menu.addAction("view_recordings",self.open_recordings_window)
                self.params_button.setMenu(self.params_Menu)
                bottom_buttons.addWidget(self.params_button)

            elif label == "Cursor":
                self.cursor_button = QPushButton("Cursor")
                self.cursor_button.setMinimumHeight(40)
                self.cursor_button.setMinimumWidth(100)
                self.cursor_button.setSizePolicy(QSizePolicy.Policy.Expanding,QSizePolicy.Policy.Fixed)
                self.cursor_button.setStyleSheet("""
                    background-color: #17a2b8;
                    color: white;
                    font-weight: bold;
                    font-size: 1rem;
                    padding:8px;
                    border-radius:8px;
                
                """)
                cursor_menu = QMenu()
                cursor_menu.addAction("Add Cursor x1", lambda: self.add_cursor("x1"))
                cursor_menu.addAction("Add Cursor x2", lambda: self.add_cursor("x2"))
                cursor_menu.addAction("Remove Cursor  x1", lambda: self.remove_cursor("x1"))
                cursor_menu.addAction("Remove cursor x2", lambda:self.remove_cursor("x2"))

                self.cursor_button.setMenu(cursor_menu)
                bottom_buttons.addWidget(self.cursor_button)


            
            else:
                btn = QPushButton(label)
                btn.setMinimumHeight(40)
                btn.setMinimumWidth(100)
                btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
                btn.setStyleSheet("""
                    background-color: #17a2b8;
                    color: white;
                    font-weight: bold;
                    font-size: 1rem;
                    padding: 8px;
                    border-radius: 8px;
                """)
                bottom_buttons.addWidget(btn)

        # Start Meas. button
        self.start_button = QPushButton("Start Meas.")
        self.start_button.setMinimumHeight(40)
        self.start_button.setMinimumWidth(100)
        self.start_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.start_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 1rem;")
        self.start_button.clicked.connect(self.start_measurement)
        bottom_buttons.addWidget(self.start_button)

        # Stop Meas. button
        self.stop_button = QPushButton("Stop Meas.")
        self.stop_button.setMinimumHeight(40)
        self.stop_button.setMinimumWidth(100)
        self.stop_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.stop_button.setStyleSheet("background-color: red; color: white; font-weight: bold; font-size: 1rem;")
        self.stop_button.clicked.connect(self.stop_measurement)
        self.stop_button.setVisible(False)
        bottom_buttons.addWidget(self.stop_button)

        self.main_layout.addLayout(bottom_buttons)
        self.setLayout(self.main_layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

    def go_back_to_dashboard(self):
        self.main_window.stacked_widget.setCurrentIndex(0)

    def on_param_selected(self, option):
        pass

    def analysis_parameter(self):
        self.popup = AnalysisParameter(self)
        self.popup.show()

    def on_param_selected(self, option):
        if option == "Display Preferences":
            self.display_preferences_dialog = DisplayPreferences(self)
            self.display_preferences_dialog.exec()
        elif option == "Input Channels":
              if self.input_channels_dialog.exec() == QDialog.accepted:
                  print("✅ Input channels updated")

    def closeEvent(self, event):
        # Save everything at shutdown
        save_settings({
            "selected_quantity": self.selected_quantity,
            "selected_fmax_hz": self.selected_fmax_hz,
            "selected_fmin_hz": self.selected_fmin_hz,
            "top_channel": self.top_channel,
            "bottom_channel": self.bottom_channel,
            "trace_mode_index": self.stacked_views.currentIndex(),
            "display_settings": self.display_settings,

            "input_channels": self.input_channels_dialog.data,
            "sample_rate": self.daq.sample_rate,
        })
        super().closeEvent(event)

    
    def get_selected_channel_sensitivities(self):
        top = self.top_channel
        bottom = self.bottom_channel

        try:
            top_sens = float(self.input_channels_dialog.data[top]["sensitivity"]) / 1000.0
        except (KeyError, IndexError, AttributeError):
            print("⚠️ Using default sensitivity: 0.1 V/g for Top Channel")
            top_sens = 0.1  # Default value in V/g

        try:
            bot_sens = float(self.input_channels_dialog.data[bottom]["sensitivity"]) / 1000.0
        except (KeyError, IndexError, AttributeError):
            print("Using default sensitivity: 0.1 V/g for Bottom Channel")
            bot_sens = 0.1  # Default value in V/g

        
        return top_sens, bot_sens
 
    

    def make_pg_plot(self, title="", xlabel="", y_label=""):
        vb = TouchViewBox()
        plot = pg.PlotWidget(viewBox=vb)
        plot.showGrid(x=True, y=True,alpha =0.3)
        plot.setTitle(title, color ='r', size ='14pt')
        plot.setLabel('left', y_label, color = 'red', size ='12pt')
        plot.setLabel('bottom', xlabel,color = "red", size ='12pt')
        plot.setAntialiasing(True)

        vb.cursor_enabled = False
        vb.cursors = {}
        return plot


    def event(self, event):
        if event.type() ==  QEvent.Type.Gesture:
            return self.gestureEvent(event)
        return super().event(event)

    def gestureEvent(self, event):
        pinch = event.gesture(Qt.GestureType.PinchGesture)
        if pinch:
            self.handle_pinch(pinch)
            return True
        return False
    
    def handle_pinch(self, pinch):
        # scale_factor = pinch.scaleFactor()
        # if scale_factor == 0:
        #     return

        # view_index = self.stacked_views.currentIndex()

        # if view_index == 1:
        #     for plot in [self.pg_waveform, self.pg_spectrum]:
        #         vb = plot.getViewBox()
        #         vb.scaleBy((1 / scale_factor, 1.0))

        # elif view_index == 2:
        #     for plot in [self.pg_waveform_top, self.pg_waveform_bottom]:
        #         vb = plot.getViewBox()
        #         vb.scaleBy((1 / scale_factor, 1.0))

        # elif view_index == 3:
        #     for plot in [self.pg_spectrum_top, self.pg_spectrum_bottom]:
        #         vb = plot.getViewBox()
        #         vb.scaleBy((1 / scale_factor, 1.0))
        pass


    def attach_focus_value_overlay(
            self, plot_widget,
            x_array, y_array,
            x_unit="s", y_unit="g",
            snap_to_peak=False, win=5):

        text_item = pg.TextItem(
            "",
            anchor=(0.5, 1.2),               # small, compact anchor
            fill=(50, 50, 50, 230),
            color=(255, 255 ,255),
            border= pg.mkPen('cyan', width =2)
        )
        text_item.setFont(QtGui.QFont("Arial", 11, QtGui.QFont.Weight.Bold))
        plot_widget.addItem(text_item)
        text_item.setZValue(2000)
        text_item.hide()

        # avoid duplicate connections
        try:
            plot_widget.scene().sigMouseClicked.disconnect()
        except:
            pass

        def handle_finger_tap(event):
            # IMPORTANT: this makes touch events work
            if event.button() != 1 and event.buttons():      # finger tap has no mouse buttons
                pass                     # allow tap
            if not len(x_array) or not len(y_array):
                return

            # convert tap → x-value
            scene_pos = event.scenePos()
            plot_item = plot_widget.getPlotItem()
            
            try:
                view_pos = plot_item.vb.mapSceneToView(scene_pos)
                x_tapped = view_pos.x()
                y_tapped = view_pos.y()
            except:
                return

            distances = np.abs(x_array - x_tapped)
            idx = int(np.argmin(distances))

            if idx < 0 or idx >= len(y_array):
                return

            if snap_to_peak and win > 0:
                start_idx = max(0, idx - win)
                end_idx = min(len(y_array), idx + win +1)

                local_y = y_array[start_idx:end_idx]
                if len(local_y) > 0:
                    local_peak = np.argmax(local_y)
                    idx = start_idx + local_peak

            x_val = x_array[idx]
            y_val = y_array[idx]

            if x_unit == "Hz":
                display_text = f"{x_val:.2f} Hz\n {y_val:.4f} {y_unit}"

            else:
                display_text = f"{x_val:.4f} {x_unit}\n {y_val:.4f} {y_unit}"

            text_item.setText(display_text)

            vb = plot_item.vb
            view_rect = vb.viewRect()

            x_min, x_max = view_rect.left(), view_rect.right()
            y_min, y_max = view_rect.bottom(), view_rect.top()

            x_margin = (x_max - x_min) * 0.08
            y_margin = (y_max - y_min) *0.1

            label_x = np.clip(x_val, x_min + x_margin, x_max + x_margin)
            label_y = np.clip(y_val + y_margin,y_min + y_margin, y_max-y_margin)

            text_item.setPos(label_x, label_y)
            text_item.show()

            QTimer.singleShot(4000, text_item.hide)

            event.accept()

        plot_widget.scene().sigMouseClicked.connect(handle_finger_tap)



    def toggle_recording(self):
        if self.record_btn.isChecked():
            self.start_recording()
            self.record_btn.setIcon(QIcon("assets/record_stop.png"))
            self.record_btn.setText("stop recording")
        else:
            self.stop_recording()
            self.record_btn.setIcon(QIcon("assets/record_start.png"))
            self.record_btn.setText("start recording")

    # def update_rpm_display(self, rpm):
    #     self.rpm_value = rpm
    #     self.rpm_label.setText(f"{rpm:.1f} RPM") 

    def start_recording(self):
    # Ensure measurement is running
        if not self.timer.isActive():
            self.start_measurement()

        timestamp = QDateTime.currentDateTime().toString("yyyyMMdd_hhmmss")
        self.session_folder = os.path.join(self.recordings_dir, timestamp)
        os.makedirs(self.session_folder, exist_ok=True)

        # 🔍 Folder check
        if not os.path.exists(self.session_folder):
            print(f"❌ Folder not created: {self.session_folder}")
            QMessageBox.critical(self, "Folder Error", "Recording folder could not be created.")
            return
        else:
            print(f"📁 Folder confirmed: {self.session_folder}")

        # 🛡️ Write permission check
        try:
            test_path = os.path.join(self.session_folder, "test_write.tmp")
            with open(test_path, 'w') as f:
                f.write("test")
            os.remove(test_path)
            print("✅ Write permission confirmed.")
        except Exception as e:
            print(f"❌ Write permission error: {e}")
            QMessageBox.critical(self, "Permission Error", "Cannot write to recording folder.")
            return

        # Setup CSV
        csv_path = os.path.join(self.session_folder, "waveform_data.csv")
        try:
            self.csv_file = open(csv_path, 'w', newline='')
            self.csv_writer = csv.writer(self.csv_file)
            self.csv_writer.writerow(["Timestamp", "Channel", "Acceleration (g)", "Velocity RMS (mm/s)", "Displacement PP (µm)", "Dominant Frequency (Hz)"])
        except Exception as e:
            print(f"❌ Failed to open CSV file: {e}")
            QMessageBox.critical(self, "CSV Error", f"Cannot write to CSV: {e}")
            return

        # 🎞️ Setup Video Writer
        screen = QApplication.primaryScreen()
        window = self.window()
        rect = window.geometry()

        print(f"🎞️ Video resolution: {rect.width()} x {rect.height()}")

        # 🧱 Check for valid window size
        if rect.width() <= 0 or rect.height() <= 0:
            print("❌ Invalid window size.")
            QMessageBox.critical(self, "Window Error", "Invalid window size. Please resize the window and try again.")
            return

        # 🧪 Try mp4v codec
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        video_path = os.path.join(self.session_folder, "recording.mp4")
        self.video_writer = cv2.VideoWriter(video_path, fourcc, 5.0, (rect.width(), rect.height()))
        print("🔎 Trying codec: mp4v")

        if not self.video_writer.isOpened():
            print("mp4v codec failed. Trying fallback: XVID")
            fourcc = cv2.VideoWriter_fourcc(*"XVID")
            video_path = os.path.join(self.session_folder, "recording.avi")
            self.video_writer = cv2.VideoWriter(video_path, fourcc, 5.0, (rect.width(), rect.height()))

            if not self.video_writer.isOpened():
                print("❌ Fallback XVID also failed.")
                QMessageBox.critical(self, "Video Error", "Video writer failed. Check codec support or permissions.")
                self.video_writer = None
                return
            else:
                print("✅ Fallback XVID worked. Recording to .avi")
        else:
            print("✅ mp4v codec worked. Recording to .mp4")

        # ✅ Ready to record
        self.recording = True
        self.frame_timer = QTimer()
        self.frame_timer.timeout.connect(self.capture_frame)
        self.frame_timer.start(200)

        print(f"🎥 Recording started: {video_path}")


    
    
    def stop_recording(self):
        self.recording =  False
        if self.csv_file:
            self.csv_file.close()
            self.csv_file = None
        if self.video_writer:
            self.video_writer.release()
        self.csv_file = None
        self.video_writer = None
        print(".Recording stopped")
    
    
    def capture_frame(self):
        if not self.recording:
            return

        screen = QApplication.primaryScreen()
        window = self.window()
        rect = window.geometry()

        image_captured = False
        if screen and window and window.isVisible() and rect.width() > 0 and rect.height() > 0:
            QApplication.processEvents()
            window.repaint()
            pixmap = screen.grabWindow(window.winId(), 0, 0, rect.width(), rect.height())
            if not pixmap.isNull():
                image = pixmap.toImage().convertToFormat(QImage.Format.Format_RGBA8888)
                ptr = image.bits()
                ptr.setsize(image.byteCount())
                rgba = np.frombuffer(ptr, np.uint8).reshape((image.height(), image.width(), 4))
                bgr = cv2.cvtColor(rgba, cv2.COLOR_RGBA2BGR)
                if self.video_writer:
                    self.video_writer.write(bgr)
                    image_captured = True
            else:
                print("⚠️ grabWindow() returned null pixmap. Frame skipped.")
        else:
            print("⚠️ Screen or window not ready for capture.")

        #  Always analyze DAQ and write data even if frame skipped
        results = self.daq.get_latest_waveform(fmax_hz=self.selected_fmax_hz, fmin_hz=self.selected_fmin_hz)

        if not results or len(results) != 2:
            print("⚠️ Incomplete DAQ results.")
            return

        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss.zzz")

        for i, result in enumerate(results):
            if not result["acceleration"].size:
                print(f"⚠️ Channel {i} has no valid acceleration data. Skipping.")
                continue

            print(f"📥 Channel {i} - acc_peak: {result['acc_peak']:.2f}g, rms_fft: {result['rms_fft']:.2f}, dom_freq: {result['dom_freq']:.2f}")

            try:
                self.csv_writer.writerow([
                    timestamp,
                    i,
                    f"{result.get('acc_peak', 0):.2f}",
                    f"{result.get('rms_fft', 0):.2f}",
                    f"{result.get('disp_pp', 0):.2f}",
                    f"{result.get('dom_freq', 0):.2f}",
                ])
            except Exception as e:
                print(f"❌ CSV write failed for channel {i}: {e}")


    def open_recordings_window(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Recordings")
        layout = QVBoxLayout(dialog)

        list_widget = QListWidget()
        recordings = [f for f in os.listdir(self.recordings_dir) if os.path.isdir(os.path.join(self.recordings_dir, f))]
        list_widget.addItems(recordings)
        layout.addWidget(list_widget)

        play_btn = QPushButton("Play Recording")
        play_btn.clicked.connect(lambda: self.play_recording(list_widget.currentItem().text()))
        layout.addWidget(play_btn)

        dialog.setLayout(layout)
        dialog.exec()

    def play_recording(self, folder_name):
        video_path =  os.path.join(self.recordings_dir, folder_name, "recording.mp4")
        if os.path.exists(video_path):
            import subprocess
            if os.name == 'posix':
                subprocess.Popen(['xdg-open', video_path])
            elif os.name == 'nt':
                os.startfile(video_path)
            elif os.name == 'mac':
                subprocess.Popen(['open', video_path])
    def take_snapshot(self):
        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd_HH-mm-ss")
        snapshot_dir = os.path.join("snapshot_data", timestamp)
        os.makedirs(snapshot_dir, exist_ok=True)

        # Take screenshot
        try:
            pixmap = self.window().grab()
            if pixmap.isNull():
                QMessageBox.warning(self, "Snapshot Error", "Screenshot failed.")
                return
            image_path = os.path.join(snapshot_dir, f"snapshot_{timestamp}.png")
            pixmap.save(image_path)
        except Exception as e:
            QMessageBox.warning(self, "Snapshot Error", f"Screenshot failed: {e}")
            return

        # Get latest or cached data
        try:
            results = self.daq.get_latest_waveform(fmax_hz=self.selected_fmax_hz, fmin_hz=self.selected_fmin_hz)
        except Exception as e:
            print("⚠️ Live read failed:", e)
            results = getattr(self, "last_snapshot_data", None)

        if not results or any(len(r["time"]) == 0 for r in results):
            QMessageBox.warning(self, "Snapshot Error", "No data available to export.")
            return

        # Prepare labels
        quantity = self.selected_quantity
        if quantity == "Acceleration":
            y_label = "Amplitude (g)"
            fft_label = "Amplitude (g)"
            wave_title = "Waveform (Time vs Acceleration)"
            fft_title = "Spectrum (Frequency vs Acceleration)"
        elif quantity == "Velocity":
            y_label = "Velocity (mm/s)"
            fft_label = "Velocity (mm/s)"
            wave_title = "Waveform (Time vs Velocity)"
            fft_title = "Spectrum (Frequency vs Velocity)"
        else:
            y_label = "Displacement (µm)"
            fft_label = "Displacement (µm)"
            wave_title = "Waveform (Time vs Displacement)"
            fft_title = "Spectrum (Frequency vs Displacement)"

        trace_mode = self.stacked_views.currentIndex()
        ch_top = self.top_channel
        ch_bot = self.bottom_channel
        file_name = f"snapshot_{quantity}_Top_Ch{ch_top}_Bottom_Ch{ch_bot}.xlsx"
        excel_path = os.path.join(snapshot_dir, file_name)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                if trace_mode == 0:
                    res = results[ch_bot]
                    df = pd.DataFrame({"Time (s)": res["time"], y_label: res[quantity.lower()]})
                    df.to_excel(writer, sheet_name="Bottom_Channel", index=False, startrow=2)
                    writer.sheets["Bottom_Channel"].cell(row=1, column=1, value=wave_title)

                elif trace_mode == 1:
                    res_top = results[ch_top]
                    res_bot = results[ch_bot]
                    df_wave = pd.DataFrame({"Time (s)": res_top["time"], y_label: res_top[quantity.lower()]})
                    df_fft = pd.DataFrame({"Frequency (Hz)": res_bot["fft_freqs"], fft_label: res_bot["fft_mags"]})
                    df_wave.to_excel(writer, sheet_name="Top_Channel", index=False, startrow=2)
                    df_fft.to_excel(writer, sheet_name="Bottom_Channel_FFT", index=False, startrow=2)
                    writer.sheets["Top_Channel"].cell(row=1, column=1, value=wave_title)
                    writer.sheets["Bottom_Channel_FFT"].cell(row=1, column=1, value=fft_title)

                elif trace_mode == 2:
                    for i, ch in enumerate([ch_top, ch_bot]):
                        res = results[ch]
                        df = pd.DataFrame({"Time (s)": res["time"], y_label: res[quantity.lower()]})
                        sheet = "Top_Channel" if i == 0 else "Bottom_Channel"
                        df.to_excel(writer, sheet_name=sheet, index=False, startrow=2)
                        writer.sheets[sheet].cell(row=1, column=1, value=wave_title)

                elif trace_mode == 3:
                    for i, ch in enumerate([ch_top, ch_bot]):
                        res = results[ch]
                        df = pd.DataFrame({"Frequency (Hz)": res["fft_freqs"], fft_label: res["fft_mags"]})
                        sheet = "Top_Channel_FFT" if i == 0 else "Bottom_Channel_FFT"
                        df.to_excel(writer, sheet_name=sheet, index=False, startrow=2)
                        writer.sheets[sheet].cell(row=1, column=1, value=fft_title)

            QMessageBox.information(self, "Snapshot Saved", f"✅ Snapshot saved in:\n{snapshot_dir}")

        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to save Excel:\n{e}")

    def build_readings_waveform_view(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # --- Top readings row
        readings_layout = QHBoxLayout()
        self.acc_input = self.create_reading_box("Acc:", "g (peak)")
        self.vel_input = self.create_reading_box("Vel:", "mm/s (RMS)")
        self.disp_input = self.create_reading_box("Disp:", "�m (P-P)")
        self.freq_input = self.create_reading_box("Freq:", "Hz")
        self.rpm_input = self.create_reading_box("RPM:", "rev/min")
        for box in [self.acc_input, self.vel_input, self.disp_input, self.freq_input, self.rpm_input]:
            readings_layout.addLayout(box["layout"])
        layout.addLayout(readings_layout)

        # --- Main waveform: Matplotlib Figure
        self.figure_waveform = Figure(figsize=(6, 3))
        self.canvas_waveform = FigureCanvas(self.figure_waveform)
        self.ax_waveform = self.figure_waveform.add_subplot(111)
        # Optional: Navigation toolbar (if you want for mouse zoom)
        self.toolbar = NavigationToolbar(self.canvas_waveform, self)
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas_waveform)

        widget.setLayout(layout)
        return widget
    def build_waveform_spectrum_view(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        self.pg_waveform = self.make_pg_plot("Waveform", "Time (s)", "Amplitude (g)")
        layout.addWidget(self.pg_waveform)
        self.pg_spectrum = self.make_pg_plot("Spectrum", "Frequency (Hz)", "Magnitude (RMS)")
        layout.addWidget(self.pg_spectrum)
        return widget
    def build_waveform_waveform_view(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        self.pg_waveform_top = self.make_pg_plot("Waveform 1", "Time (s)", "Amplitude (g)")
        layout.addWidget(self.pg_waveform_top)
        self.pg_waveform_bottom = self.make_pg_plot("Waveform 2", "Time (s)", "Amplitude (g)")
        layout.addWidget(self.pg_waveform_bottom)
        return widget
    def build_spectrum_spectrum_view(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        self.pg_spectrum_top = self.make_pg_plot("Spectrum 1", "Frequency (Hz)", "Magnitude (RMS)")
        layout.addWidget(self.pg_spectrum_top)
        self.pg_spectrum_bottom = self.make_pg_plot("Spectrum 2", "Frequency (Hz)", "Magnitude (RMS)")
        layout.addWidget(self.pg_spectrum_bottom)
        return widget


    def build_reading_reading_view(self):
        widget =  QWidget()
        layout = QVBoxLayout(widget)

        readings_layout = QHBoxLayout()
        self.acc_input0 = self.create_reading_box("Acc:", "g (peak)")
        self.vel_input0 = self.create_reading_box("Vel:", "mm/s (RMS)")
        self.disp_input0 = self.create_reading_box("Disp:", "µm (P-P)")
        self.freq_input0 = self.create_reading_box("Freq:", "Hz")

        for box in [self.acc_input0, self.vel_input0, self.disp_input0, self.freq_input0]:
            readings_layout.addLayout(box["layout"])
        layout.addLayout(readings_layout)

        readings_layout1 = QHBoxLayout()
        self.acc_input1 = self.create_reading_box("Acc:","g (peak)")
        self.vel_input1 = self.create_reading_box("Vel:", "mm/s(RMS)")
        self.disp_input1 = self.create_reading_box("Disp:","µm (P-P)")
        self.freq_input1 = self.create_reading_box("Freq:","Hz")

        for box1 in [self.acc_input1,self.vel_input1,self.disp_input1,self.freq_input1]:
            readings_layout1.addLayout(box1["layout"])
        layout.addLayout(readings_layout1)

        return widget
    
    def build_readings_spectrum_view(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        readings_layout = QHBoxLayout()
        self.acc_input2 = self.create_reading_box("Acc:","g (peak)")
        self.vel_input2 = self.create_reading_box("Vel:", "mm/s (RMS)")
        self.disp_input2 = self.create_reading_box("Disp:", "�m (P-P)")
        self.freq_input2 = self.create_reading_box("Freq:", "Hz")
        self.rpm_input1 = self.create_reading_box("RPM:", "rev/min")

        for box in [self.acc_input2, self.vel_input2, self.disp_input2, self.freq_input2, self.rpm_input1]:
            readings_layout.addLayout(box["layout"])
        layout.addLayout(readings_layout)


        self.pg_readings_spectrum = self.make_pg_plot("spectrum","Frequency (Hz)", "Magnitude (RMS)")
        layout.addWidget(self.pg_readings_spectrum)
        return widget


    def create_reading_box(self, label_text, unit_text):
        layout = QVBoxLayout()
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; font-size: 1rem;")
        input_field = QLineEdit("0")
        input_field.setAlignment(Qt.AlignmentFlag.AlignCenter)
        input_field.setReadOnly(True)
        input_field.setFixedWidth(80)
        unit_label = QLabel(unit_text)
        unit_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unit_label.setStyleSheet("font-size: 0.9rem; color: #666;")
        layout.addWidget(label)
        layout.addWidget(input_field)
        layout.addWidget(unit_label)
        return {"layout": layout, "input": input_field,"unit": unit_label}

    def start_clock(self):
        self.clock = QTimer()
        self.clock.timeout.connect(self.update_time)
        self.clock.start(1000)
        self.update_time()

    def update_time(self):
        now = QDateTime.currentDateTime()
        self.time_label.setText(now.toString("HH:mm:ss"))
        
    def switch_trace_mode(self, index):
        self.stacked_views.setCurrentIndex(index)


    def open_trace_window_settings(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QComboBox, QDialogButtonBox
        from PyQt6.QtCore import Qt

        # Show current trace mode label
        trace_mode_titles = {
            0: "Readings + Waveform",
            1: "Waveform + Spectrum",
            2: "Waveform + Waveform",
            3: "Spectrum + Spectrum",
            4: "Readings + Readings",
            5:"Readings + Spectrum"
        }
        selected_trace_mode = self.stacked_views.currentIndex()
        trace_mode_label = trace_mode_titles.get(selected_trace_mode, "Unknown")

        dialog = QDialog(self)
        dialog.setWindowTitle("Trace Window Settings")
        layout = QVBoxLayout(dialog)

        # Trace mode display
        trace_mode_display = QLabel(f"Current Trace Mode: {trace_mode_label}")
        trace_mode_display.setAlignment(Qt.AlignmentFlag.AlignCenter)
        trace_mode_display.setStyleSheet("font-size: 1.2rem; font-weight: bold; color: #005cbf;")
        layout.addWidget(trace_mode_display)

        # Top trace channel selection
        label_top = QLabel("Select Top Trace Channel:")
        combo_top = QComboBox()
        combo_top.addItems(["Ch0", "Ch1", "Ch2", "Ch3", "Ch4", "Ch5"])
        combo_top.setCurrentIndex(self.top_channel)

        # Bottom trace channel selection
        label_bottom = QLabel("Select Bottom Trace Channel:")
        combo_bottom = QComboBox()
        combo_bottom.addItems(["Ch0", "Ch1", "Ch2", "Ch3", "Ch4", "Ch5"])
        combo_bottom.setCurrentIndex(self.bottom_channel)

        layout.addWidget(label_top)
        layout.addWidget(combo_top)
        layout.addWidget(label_bottom)
        layout.addWidget(combo_bottom)

        # OK / Cancel buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)

        def apply_selection():
            self.top_channel = combo_top.currentIndex()
            self.bottom_channel = combo_bottom.currentIndex()

            save_settings({
                "selected_quantity": self.selected_quantity,
                "selected_fmax_hz": self.selected_fmax_hz,
                "selected_fmin_hz": self.selected_fmin_hz,
                "top_channel": self.top_channel,
                "bottom_channel": self.bottom_channel,
                "trace_mode_index": self.stacked_views.currentIndex(),
                "display_settings": self.display_settings,
                "input_channels": self.input_channels_dialog.data,
                "sample_rate": self.daq.sample_rate
            })
            dialog.accept()

        buttons.accepted.connect(apply_selection)
        buttons.rejected.connect(dialog.reject)

        dialog.exec()

    # def update_rpm_display(self, rpm):
    #     self.rpm_input["input"].setText(f"{rpm:.0f}")


   
    def start_measurement(self):
        
        self.daq.start_acquisition()
        self.timer.start(self.update_interval_ms)
        self.start_button.setVisible(False)
        self.stop_button.setVisible(True)


    def stop_measurement(self):
        self.daq.stop_scan()
        self.timer.stop()
        #self.is_running = False
        self.start_button.setVisible(True)
        self.stop_button.setVisible(False)
        print("Measurement stopped.")

    def add_cursor(self, name):
        if name in self.cursors:
            print(f"[Cursor] {name} already exists")
            return

        view_index = self.stacked_views.currentIndex()
        if view_index == 1:
            plots = [self.pg_spectrum]
        elif view_index == 3:
            plots = [self.pg_spectrum_top, self.pg_spectrum_bottom]
        elif view_index == 5:
            plots = [self.pg_readings_spectrum]
        else:
            return

        cursor_color = 'yellow' if name.lower() == "x1" else 'cyan'

        cursors = []
        labels = []

        for plot in plots:
            vb = plot.getViewBox()
            current_x_range = vb.viewRange()[0]
            current_y_range = vb.viewRange()[1]
            
            plot_id = id(plot)
            self.locked_ranges[plot_id] = (current_x_range, current_y_range)

            cursor = DraggableCursor(
                parent_plot = plot,
                angle=90,
                movable = True,
                pen = pg.mkPen(cursor_color, width =2, style = Qt.PenStyle.DashLine),
                name=name
            )
            cursor.setZValue(1000)

            x_center = (current_x_range[0] + current_x_range[1]) / 2
            cursor.setValue(x_center)

            plot.addItem(cursor)

            label = pg.TextItem(
                "",
                anchor = (0.5,0),
                fill = (50,50,50,200),
                color = cursor_color,
                border = pg.mkPen(cursor_color, width =2)
            )
            label.setFont(QtGui.QFont("Arial",10,QtGui.QFont.Weight.Bold))
            label.setZValue(1001)
            plot.addItem(label)

            # Store references
            cursors.append(cursor)
            labels.append(label)

            # Connect signal to update label
            cursor.sigPositionChanged.connect(partial(self.update_cursor_label,cursor,label,plot))

            vb.cursor_enabled = True
            vb.cursors[name] = cursors

        self.cursors[name] = cursors
        self.cursor_value_labels[name] =  labels

        self.cursor_enabled = True

        for cursor, label, plot in zip(cursors, labels, plots):
            self.update_cursor_label(cursor, label, plot)

        self.update_cursor_delta()

        print(f" Cursor {name} added to {len(plots)} plot(s)")
    # def update_spectrum_label(self, cursor, label, plot):
    #     pos = cursor.value()
    #     data_items = plot.listDataItems()

    #     if not data_items:
    #         label.setText("")
    #         return

    #     x_data = data_items[0].xData
    #     y_data = data_items[0].yData

    #     if x_data is None or y_data is None or len(x_data) == 0:
    #         label.setText("")
    #         return

    #     # Find the closest index
    #     idx = np.abs(x_data - pos).argmin()
    #     x_val = x_data[idx]
    #     y_val = y_data[idx]

    #     unit_x = "Hz"
    #     unit_y = plot.getAxis('left').labelText or "unit"

    #     label.setText(f"{cursor.name()}: {x_val:.2f}{unit_x}, {y_val:.2f}{unit_y}")

    #     # Keep label at the top of visible area
    #     vb = plot.getViewBox()
    #     xr, yr = vb.viewRange()
    #     label.setPos(x_val, yr[1])  # Y at top of current view

    #     # Prevent auto-scale
    #     vb.setXRange(*xr, padding=0)
    #     vb.setYRange(*yr, padding=0)
    def update_cursor_label(self, cursor, label, plot):
        """Update individual cursor label without affecting plot scaling"""
        # Get current view range and LOCK IT
        vb = plot.getViewBox()
        plot_id = id(plot)
        
        # Get locked range if it exists
        if plot_id in self.locked_ranges:
            locked_x, locked_y = self.locked_ranges[plot_id]
        else:
            locked_x, locked_y = vb.viewRange()[0], vb.viewRange()[1]
            self.locked_ranges[plot_id] = (locked_x, locked_y)
        
        # Get cursor position
        x_pos = cursor.value()
        
        # Get data from plot
        data_items = plot.listDataItems()
        if not data_items:
            label.setText("")
            return

        x_data = data_items[0].xData
        y_data = data_items[0].yData

        if x_data is None or y_data is None or len(x_data) == 0:
            label.setText("")
            return

        # Find nearest data point
        idx = np.abs(x_data - x_pos).argmin()
        x_val = x_data[idx]
        y_val = y_data[idx]

        # Get units from axis labels
        x_label = plot.getAxis('bottom').labelText or "X"
        y_label = plot.getAxis('left').labelText or "Y"
        
        # Determine units
        if "Hz" in x_label or "Frequency" in x_label:
            unit_x = "Hz"
        else:
            unit_x = "s"
        
        # Extract unit from y_label (e.g., "g (RMS)" -> "g")
        if "(" in y_label:
            unit_y = y_label.split("(")[0].strip()
        else:
            unit_y = y_label

        # Format label text
        label_text = f"{cursor.name()}\n{x_val:.2f} {unit_x}\n{y_val:.4f} {unit_y}"
        label.setText(label_text)

        # Position label at top of cursor, within view
        label_y = locked_y[1] - (locked_y[1] - locked_y[0]) * 0.05  # Near top
        label.setPos(x_val, label_y)

        # CRITICAL: Restore locked range to prevent scaling
        vb.setRange(xRange=locked_x, yRange=locked_y, padding=0)
        
        # Update delta display
        self.update_cursor_delta()

    def remove_cursor(self, name):
        view_index = self.stacked_views.currentIndex()

        if view_index == 1:
            plots = [self.pg_spectrum]

        elif view_index == 2:
            plots = [self.pg_waveform_top, self.pg_waveform_bottom]
        elif view_index == 3:
            plots = [self.pg_spectrum_top, self.pg_spectrum_bottom]
        elif view_index == 5:
            plots = [self.pg_readings_spectrum]
        else:
            return
        
        if name not in self.cursors:
             return

        for plot,cursor, label in zip(plots,self.cursors[name],self.cursor_value_labels[name]):
            plot.removeItem(cursor)
            plot.removeItem(label)

            vb = plot.getViewBox()
            if name in vb.cursors:
                del vb.cursors[name]

            if not vb.cursors:
                vb.cursor_enabled = False

        del self.cursors[name]
        del self.cursor_value_labels[name]

        if not self.cursors:
            self.cursor_enabled = False
        self.update_cursor_delta()

    def update_cursor_delta(self):
        if "x1" not in self.cursors or "x2" not in self.cursors:
            self.cursor_info_label.setText("Cursor Info: ")
            return

        view_index = self.stacked_views.currentIndex()
        
        unit_x = "s" if view_index in (0, 2) else "Hz"

        try:
            x1 = self.cursors["x1"][0].value()
            x2 = self.cursors["x2"][0].value()
        except Exception as e:
            self.cursor_info_label.setText("Cursor Info: Error reading cursor values")
            return

        delta_x = abs(x2 - x1)

        if unit_x == "s" and delta_x > 0:
            freq_hz = 1.0 / delta_x
            freq_info = f" | Frequency: {freq_hz:.2f} Hz"
        else:
            freq_info = ""

        # Pick plots based on current view
        if view_index == 1:
            plots = [self.pg_waveform]
        elif view_index == 2:
            plots = [self.pg_waveform_top, self.pg_waveform_bottom]
        elif view_index == 3:
            plots = [self.pg_spectrum_top, self.pg_spectrum_bottom]
        elif view_index == 5:
            plots = [self.pg_readings_spectrum]
        else:
            self.cursor_info_label.setText("Cursor Info: Unsupported View")
            return

        y_vals = []
        for plot in plots:
            data_items = plot.listDataItems()
            if not data_items:
                y_vals.append(("?", "?"))
                continue

            x_array = data_items[0].xData
            y_array = data_items[0].yData

            def nearest_y(x_val):
                idx = np.argmin(np.abs(x_array - x_val))
                return y_array[idx]

            y1 = nearest_y(x1)
            y2 = nearest_y(x2)
            y_vals.append((y1, y2))

        # Compose label text
        label_lines = [
            f"X1: {x1:.2f} {unit_x}, X2: {x2:.2f} {unit_x}, ?X: {delta_x:.2f} {unit_x}"
        ]

        for i, (y1, y2) in enumerate(y_vals):
            delta_y = abs(y2 - y1)
            unit_y = plots[i].getAxis('left').labelText or "unit"
            line = f" | [{i}] Y1: {y1:.2f}, Y2: {y2:.2f}, ?Y: {delta_y:.2f} {unit_y}"
            label_lines.append(line)

        self.cursor_info_label.setText("".join(label_lines))

    def update_plot_with_zoom_preservation(self, plot, x_data, y_data,
                                       pen_color='cyan',
                                       x_label="Frequency (Hz)",
                                       y_label=None):

        vb = plot.getViewBox()

        if vb.user_has_zoomed:
            vb.enableAutoRange(enable=False)
            saved_range = vb.manual_zoom_range
        else:
            saved_range = None

        plot.clear()
        plot.plot(x_data, y_data, pen=pg.mkPen(pen_color, width=2))

        # ✅ RE-APPLY LABELS AFTER CLEAR
        plot.setLabel('bottom', x_label, color='w', size='12pt')
        if y_label:
            plot.setLabel('left', y_label, color='w', size='12pt')

        if saved_range:
            vb.setRange(xRange=saved_range[0], yRange=saved_range[1], padding=0)

    def reset_all_zoom(self):

        view_index =  self.stacked_views.currentIndex()

        if view_index ==1:
            self.pg_waveform.getViewBox().reset_zoom()
            self.pg_spectrum.getViewBox().reset_zoom()

        elif view_index == 2:
            self.pg_waveform_top.getViewBox().reset_zoom()
            self.pg_waveform_bottom.getViewBox().reset_zoom()
        elif view_index == 3:
            self.pg_spectrum_top.getViewBox().reset_zoom()
            self.pg_spectrum_bottom.getViewBox().reset_zoom()
        elif view_index == 5:
            self.pg_readings_spectrum.getViewBox().reset_zoom()

    


    from math import sqrt

# ────────────────────────────────────────────────────────────────────────
#  Helpers – only define once in your file
# ────────────────────────────────────────────────────────────────────────
    def spec_factor(self,base_kind: str, want: str) -> float:
        """
        Convert a value expressed as Peak / RMS / Peak-Peak (base_kind)
        to the requested spectrum type (want).
        """
        tbl = {
            "peak": {"Peak": 1, "RMS": 1 / sqrt(2), "Peak-Peak": 2},
            "rms":  {"RMS": 1, "Peak": sqrt(2),     "Peak-Peak": 2 * sqrt(2)},
            "pp":   {"Peak-Peak": 1, "Peak": 0.5,   "RMS": 0.5 / sqrt(2)},
        }
        return tbl[base_kind][want]

    _UNIT_FACTOR = {
        "Acceleration": {"g": 1, "m/s²": 9.80665, "mm/s²": 9806.65,
                        "cm/s²": 980.665, "in/s²": 386.089},
        "Velocity":     {"mm/s": 1, "cm/s": 0.1, "m/s": 0.001, "in/s": 0.0393701},
        "Displacement": {"µm": 1, "μm": 1, "mm": 0.001, "cm": 0.0001,
                        "m": 1e-6, "in": 0.0000393701},
    }
    # ────────────────────────────────────────────────────────────────────────

    def update_plot(self):
        # ─── Get newest data ────────────────────────────────────────────────
        top_sens, bot_sens = self.get_selected_channel_sensitivities()
        results = self.daq.get_latest_waveform(
            fmax_hz=self.selected_fmax_hz,
            fmin_hz=self.selected_fmin_hz,
            sensitivities=[top_sens, bot_sens],
        )
        if not results:
            return

        channel_to_result = {
            ch: res for ch, res in zip(self.daq.channel, results)
        }
        top_result    = channel_to_result.get(self.top_channel,    self.daq.empty_result())
        bottom_result = channel_to_result.get(self.bottom_channel, self.daq.empty_result())
        if self.top_channel == self.bottom_channel:
            bottom_result = top_result

        # ─── Display-preference pairs (unit + spec) ─────────────────────────
        acc_spec = self.display_settings.get("Acceleration Spectrum Type",  "Peak")
        acc_unit = self.display_settings.get("Acceleration Engineering Unit", "g")
        vel_spec = self.display_settings.get("Velocity Spectrum Type",      "RMS")
        vel_unit = self.display_settings.get("Velocity Engineering Unit",   "mm/s")
        disp_spec = self.display_settings.get("Displacement Spectrum Type", "Peak-Peak")
        disp_unit = self.display_settings.get("Displacement Engineering Unit", "µm")

        quantity = self.selected_quantity
        unit = self.display_settings.get(f"{quantity} Engineering Unit")
        spec_type = self.display_settings.get(f"{quantity} Spectrum Type")

        OVERALL_KEYS = {
            "Acceleration": {
                "Peak":        "acc_peak",
                "RMS":         "acceleration_rms",
                "Peak-Peak":   "acceleration_ptps",
            },
            "Velocity": {
                "Peak":        "velocity_peak",
                "RMS":         "velocity_rms",
                "Peak-Peak":   "velocity_ptps",
            },
            "Displacement": {
                "Peak":        "displacement_peak",
                "RMS":         "displacement_rms",
                "Peak-Peak":   "displacement_ptps",
            },
        }

        def get_overall_vibration_values(result, quantity, spec_type):
            key = OVERALL_KEYS[quantity][spec_type]
            return result.get(key, 0.0)

        # ─── Get overall values for readings ────────────────────────────────
        acc_val = get_overall_vibration_values(top_result, "Acceleration", acc_spec)
        vel_val = get_overall_vibration_values(top_result, "Velocity", vel_spec)
        disp_val = get_overall_vibration_values(top_result, "Displacement", disp_spec)

        acc_val *= self._UNIT_FACTOR["Acceleration"][acc_unit]
        vel_val *= self._UNIT_FACTOR["Velocity"][vel_unit]
        disp_val *= self._UNIT_FACTOR["Displacement"][disp_unit]

        acce_val  = get_overall_vibration_values(bottom_result, "Acceleration", acc_spec)
        vel_val1 = get_overall_vibration_values(bottom_result, "Velocity", vel_spec)
        disp_val1 = get_overall_vibration_values(bottom_result, "Displacement", disp_spec)

        acce_val *= self._UNIT_FACTOR["Acceleration"][acc_unit]
        vel_val1 *= self._UNIT_FACTOR["Velocity"][vel_unit]
        disp_val1 *= self._UNIT_FACTOR["Displacement"][disp_unit]

        # ─── Pick current quantity for waveform / spectrum view ─────────────
        if quantity == "Velocity":
            y_top = top_result["velocity"]
            y_bot = bottom_result["velocity"]
            fft_top = top_result["fft_mags_vel"]
            fft_bot = bottom_result["fft_mags_vel"]
            spec_choice, unit_choice = vel_spec, vel_unit
        elif quantity == "Displacement":
            y_top = top_result["displacement"]
            y_bot = bottom_result["displacement"]
            fft_top = top_result["fft_mags_disp"]
            fft_bot = bottom_result["fft_mags_disp"]
            spec_choice, unit_choice = disp_spec, disp_unit
        else:
            y_top = top_result["acceleration"]
            y_bot = bottom_result["acceleration"]
            fft_top = top_result["fft_mags"]
            fft_bot = bottom_result["fft_mags"]
            spec_choice, unit_choice = acc_spec, acc_unit

        # ─── Apply unit/spec scaling to waveform & spectrum data ────────────
        unit_factor = self._UNIT_FACTOR[quantity][unit_choice]
        spec_scaler_fft = self.spec_factor("peak", spec_choice)

        y_top = y_top * unit_factor
        y_bot = y_bot * unit_factor
        mag_top = fft_top * spec_scaler_fft * unit_factor
        mag_bot = fft_bot * spec_scaler_fft * unit_factor
        self.freqs = top_result["frequencies"]

        # ─── GUI update according to current stacked view ──────────────────
        view = self.stacked_views.currentIndex()
        t_vec = top_result["time"]
        y_label_wave = unit_choice
        y_label_spec = f"{unit_choice} ({spec_choice})"

        if view == 0:  # Readings + Waveform
            self.acc_input["input"].setText(f"{acc_val:.2f}")
            self.vel_input["input"].setText(f"{vel_val:.2f}")
            self.disp_input["input"].setText(f"{disp_val:.2f}")

            self.ax_waveform.clear()
            self.ax_waveform.plot(t_vec, y_bot)
            self.ax_waveform.set_xlabel("Time (s)")
            self.ax_waveform.set_ylabel(y_label_wave)
            self.ax_waveform.grid(True)
            self.canvas_waveform.draw()

        elif view == 1:  # Waveform + Spectrum
            # Waveform
            self.pg_waveform.clear()
            self.pg_waveform.plot(t_vec, y_top, pen='b')
            self.pg_waveform.setLabel('bottom', 'Time (s)', color='w', size='12pt')
            self.pg_waveform.setLabel('left', y_label_wave, color='w', size='12pt')
            self.attach_focus_value_overlay(self.pg_waveform, t_vec, y_top, x_unit="s", y_unit=y_label_wave, snap_to_peak=False)
            
            # Spectrum
            self.pg_spectrum.clear()
            self.pg_spectrum.plot(self.freqs, mag_bot, pen='r')
            self.pg_spectrum.setLabel('bottom', 'Frequency (Hz)', color='w', size='12pt')
            self.pg_spectrum.setLabel('left', y_label_spec, color='w', size='12pt')
            self.attach_focus_value_overlay(self.pg_spectrum, self.freqs, mag_bot, "Hz", y_label_spec, True, win=10)

        elif view == 2:  # Waveform + Waveform
            # Top waveform
            self.pg_waveform_top.clear()
            self.pg_waveform_top.plot(t_vec, y_top, pen='b')
            self.pg_waveform_top.setLabel('bottom', 'Time (s)', color='red', size='12pt')
            self.pg_waveform_top.setLabel('left', y_label_wave, color='red', size='12pt')
            self.attach_focus_value_overlay(self.pg_waveform_top, t_vec, y_top, "s", y_label_wave, True)

            # Bottom waveform
            self.pg_waveform_bottom.clear()
            self.pg_waveform_bottom.plot(t_vec, y_bot, pen='g')
            self.pg_waveform_bottom.setLabel('bottom', 'Time (s)', color='red', size='12pt')
            self.pg_waveform_bottom.setLabel('left', y_label_wave, color='red', size='12pt')
            self.attach_focus_value_overlay(self.pg_waveform_bottom, t_vec, y_bot, "s", y_label_wave, True)

        elif view == 3:  # Spectrum + Spectrum
            # Top spectrum
            self.pg_spectrum_top.clear()
            self.pg_spectrum_top.plot(self.freqs, mag_top, pen='b')
            self.pg_spectrum_top.setLabel('bottom', 'Frequency (Hz)', color='red', size='12pt')
            self.pg_spectrum_top.setLabel('left', y_label_spec, color='red', size='12pt')
            self.attach_focus_value_overlay(self.pg_spectrum_top, self.freqs, mag_top, "Hz", y_label_spec, True, win=10)

            # Bottom spectrum
            self.pg_spectrum_bottom.clear()
            self.pg_spectrum_bottom.plot(self.freqs, mag_bot, pen='g')
            self.pg_spectrum_bottom.setLabel('bottom', 'Frequency (Hz)', color='red', size='12pt')
            self.pg_spectrum_bottom.setLabel('left', y_label_spec, color='red', size='12pt')
            self.attach_focus_value_overlay(self.pg_spectrum_bottom, self.freqs, mag_bot, "Hz", y_label_spec, True, win=10)

        elif view == 4:  # Readings + Readings
            self.acc_input0["input"].setText(f"{acc_val:.2f}")
            self.vel_input0["input"].setText(f"{vel_val:.2f}")
            self.disp_input0["input"].setText(f"{disp_val:.2f}")

            self.acc_input1["input"].setText(f"{acce_val:.2f}")
            self.vel_input1["input"].setText(f"{vel_val1:.2f}")
            self.disp_input1["input"].setText(f"{disp_val1:.2f}")

        elif view == 5:  # Readings + Spectrum
            self.acc_input2["input"].setText(f"{acce_val:.2f}")
            self.vel_input2["input"].setText(f"{vel_val1:.2f}")
            self.disp_input2["input"].setText(f"{disp_val1:.2f}")
            
            # Spectrum
            self.pg_readings_spectrum.clear()
            self.pg_readings_spectrum.plot(self.freqs, mag_top, pen='b')
            self.pg_readings_spectrum.setLabel('bottom', 'Frequency (Hz)', color='red', size='12pt')
            self.pg_readings_spectrum.setLabel('left', y_label_spec, color='red', size='12pt')
            self.attach_focus_value_overlay(self.pg_readings_spectrum, self.freqs, mag_top, "Hz", y_label_spec, True, win=10)

        # Cache latest for snapshot
        self.last_snapshot_data = results